


import openpyxl as opyxl
import utility
from ortools.sat.python import cp_model



excel_file = "./input_data.xlsx"

# Read attendees names
attendees = utility.read_attendees()


wb = opyxl.load_workbook(excel_file, data_only=True)

# Read colors legend
sh = wb["Settings"]
color_holidays = sh["AS42"].fill.start_color.index[2:]
color_workday = sh["AS43"].fill.start_color.index[2:]
color_error = sh["AS44"].fill.start_color.index[2:]


# Read global constraints 
sh = wb["Global constraints"]

# I don't know if they are needed
month = sh["A2"].value.split()[0]
year = sh["A2"].value.split()[1]

# Creation of time slots
time_slots = []
for row in sh.iter_rows(min_row=5, max_row=sh.max_row, min_col=2, max_col=sh.max_column):
    for cell in row:
        col = cell.fill.start_color.index[2:]
        if col == color_workday and cell.value != "X":
            time_slots.append(sh.cell(row=3, column=cell.column).value + " / " + sh.cell(row=cell.row, column=1).value) # così è un casino, ma non so come fare meglio
        if cell.value is not None and cell.value.capitalize() != "X":
            raise ValueError(f"Error in the Global Constraints - day {sh.cell(row=3, column=cell.column).value}, timeslot {sh.cell(row=cell.row, column=1).value}")




# Read meetings and attendees
sh = wb["Groups compositions"]
meet_attend = {}
for row in sh.iter_rows(min_row=4, max_row=sh.max_row, min_col=2, max_col=2):
    for cell in row:
        meet_attend[cell.value] = []
        for col in range(3, sh.max_column + 1):
            if sh.cell(row=cell.row, column=col).value is not None:
                meet_attend[cell.value].append(sh.cell(row=cell.row, column=col).value)


# Read attendees constraints
sh = wb["Attendees constraints"]
attendee_constraints = {}
for row in sh.iter_rows(min_row=4, max_row=sh.max_row, min_col=3, max_col=sh.max_column):
    attendee_constraints[sh.cell(row=row[0].row, column=2).value] = []
    for cell in row:
        if cell.value is not None and cell.value.capitalize() == "X":
            attendee_constraints[sh.cell(row=cell.row, column=2).value].append(sh.cell(row=2, column=cell.column).value)

# wb.close()


# Data post-processing

# Index of time slots for each participant based on their constraints
ind_per_part = {}
for nome, date in attendee_constraints.items():
    ind_per_part[nome] = []
    for data in date:
        ind_per_part[nome].extend([i for i, slot in enumerate(time_slots) if slot.startswith(data)])


model = cp_model.CpModel()


# Creation of DVs
meeting_slot = {}
for meeting in meet_attend.keys():
    for time_slot in time_slots:
        meeting_slot[(meeting, time_slot)] = model.NewBoolVar(f'meeting_{meeting}_timeslot_{time_slot}')

# # Constraint 1: Each meeting must be assigned to exactly one slot
for meeting in meet_attend.keys():
    model.Add(sum(meeting_slot[(meeting, time_slot)] for time_slot in time_slots) == 1)

# Constraint 2: A person cannot attend two meetings in the same slot
for attendee in attendees:
    for time_slot in time_slots:
        model.Add(sum(meeting_slot[(meeting, time_slot)] for meeting in meet_attend
                        if attendee in meet_attend[meeting]) <= 1)
        

# Constraint 3: Personal constraints
for meeting in meet_attend:
    for attendee in meet_attend[meeting]:
        if len(ind_per_part[attendee]) > 0:
            for j in ind_per_part[attendee]:
                model.Add(meeting_slot[(meeting, time_slots[j])] == 0)

# Run the solver
solver = cp_model.CpSolver()
status = solver.Solve(model)


# Write the solution in the excel file
wb = opyxl.load_workbook(excel_file)
result_sheet = "Meetings arrangement"
# Create Meetings arrangement sheet
if result_sheet in wb.sheetnames:
    del wb[result_sheet]

# Create the sheet for the results
wb.create_sheet(result_sheet)
sh = wb[result_sheet]

# Merge two columns 
sh.merge_cells('B2:D2')
sh.merge_cells('C3:D3')
sh.cell(row=2, column=2).value = "Meetings arrangement"
sh.cell(row=3, column=2).value = "Meeting"
sh.cell(row=3, column=3).value = "Time slot"

# Write the solution in the excel file
if status == cp_model.OPTIMAL:
    row = 4
    for meeting in meet_attend.keys():
        for time_slot in time_slots:
            if solver.Value(meeting_slot[(meeting, time_slot)]):
                sh.cell(row=row, column=2).value = meeting
                sh.cell(row=row, column=3).value = time_slot
                row += 1
else:
    sh.cell(row=4, column=2).value = "No optimal solution found."

# Save the excel file if it is not open, otherwise close it and save it
while True:
    try:
        wb.save(excel_file)
        wb.close()
        break
    except PermissionError:
        input(f"Please close the file {excel_file} and press Enter to continue...")

# if status == cp_model.OPTIMAL:
#     print('Meetings arrangement found:')
#     for meeting in meet_attend.keys():
#         for time_slot in time_slots:
#             if solver.Value(meeting_slot[(meeting, time_slot)]):
#                 print(f'{meeting} on {time_slot}')
# else:
#     print('No optimal solution found.')
