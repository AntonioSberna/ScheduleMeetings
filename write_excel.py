

import os 
import openpyxl as opyxl
from datetime import datetime
from dateutil.relativedelta import relativedelta
import utility
# to install: holidays, openpyxl




# Next month
next_month = datetime.now() + relativedelta(months=1)


# Slot ranges
slot_ranges = [[7, 12], [14, 16]]
slot_duration = 1
time_slots = utility.create_timeslots(slot_ranges, slot_duration)

# Weekly day off
weekly_day_off = {"Friday", "Saturday", "Sunday"}



# Create a attendees sheet if not exists
if not os.path.exists("attendees.xlsx"):
    wb = opyxl.Workbook()
    ws = wb.active
    ws.title = "attendees"
    ws["C2"] = "Attendees"
    ws["C2"].font = opyxl.styles.Font(bold=True)
    wb.save("attendees.xlsx")
    wb.close()



# Excel
wb = opyxl.Workbook()

# Global constraints
ws = wb.active
ws.title = "Global constraints"

# Insert month
ws.cell(row=2, column=1).value = f"{next_month.strftime('%B')} {next_month.year}"
ws.cell(row=2, column=1).font = opyxl.styles.Font(bold=True)


# Days in month
start_row = 4
utility.write_days_holidays(ws, next_month, len(time_slots), weekly_day_off, start_row = start_row)


# Insert timeslots
ws.column_dimensions["A"].width = 11.5
for i, time_slot in enumerate(time_slots, start=1):
    ws.cell(row=start_row + i, column=1).value = f"{time_slot} - {time_slots[i] if i < len(time_slots) else f"{slot_ranges[-1][-1]}:00"}" # can be improved the last slot's end




# Team dayoff
ws = wb.create_sheet(title="Attendees constraints")
attendees = utility.read_attendees()

if attendees is []:
    raise ValueError("No attendees found in the attendees.xlsx file")

start_row = 4
# Insert names
ws["B3"] = "Names"
ws.cell(row=start_row-1, column=2).font = opyxl.styles.Font(bold=True, underline="single")
ws.cell(row=start_row-1, column=2).alignment = opyxl.styles.Alignment(horizontal="center", vertical="center")
ws.column_dimensions[opyxl.utils.get_column_letter(2)].width = 11.5



# Day of the month
utility.write_days_holidays(ws, next_month, len(attendees), weekly_day_off, start_row = start_row-1, start_col = 2)

# Insert attendees' names
for i, attendee in enumerate(attendees):
    ws[f"B{i+start_row}"] = attendee
    ws.cell(row=i+start_row, column=2).font = opyxl.styles.Font(bold=True)




# Groups compositions
ws = wb.create_sheet(title="Groups compositions")
ws["B3"] = "Groups"
ws.cell(row=start_row-1, column=2).font = opyxl.styles.Font(bold=True, underline="single")
ws.cell(row=start_row-1, column=2).alignment = opyxl.styles.Alignment(horizontal="center", vertical="center")
ws.column_dimensions[opyxl.utils.get_column_letter(2)].width = 11.5


# Settings sheet
ws = wb.create_sheet(title="Settings")
# Color samples (da sistemare)
ws["AS42"].fill = opyxl.styles.PatternFill(start_color="9a0000", end_color="9a0000", fill_type="solid")
ws["AS43"].fill = opyxl.styles.PatternFill(start_color="009a00", end_color="009a00", fill_type="solid")
ws["AS44"].fill = opyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
ws.sheet_state = 'hidden'

# Salva il file
# Add check if file already exists
wb.save("./input_data.xlsx")
wb.close()




