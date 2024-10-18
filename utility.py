

import openpyxl as opyxl
import holidays
from datetime import datetime
import calendar



def write_days_holidays(ws, next_month, n_rows, weekly_day_off, **kwargs):

    color_holidays = kwargs.get("color_holidays", "9a0000")
    color_workday = kwargs.get("color_workday", "009a00")
    column_width = kwargs.get("column_width", 6.75)
    start_row = kwargs.get("start_row", 4)
    start_col = kwargs.get("start_col", 1)
    
    days_in_month = calendar.monthrange(next_month.year, next_month.month)[1]

    for day in range(1, days_in_month + 1):
        col = day+start_col
        ws.column_dimensions[opyxl.utils.get_column_letter(col)].width = column_width

        # Day of the month
        ws.cell(row=start_row-1, column=col).value = f"{day} {next_month.strftime('%b')}"
        ws.cell(row=start_row-1, column=col).alignment = opyxl.styles.Alignment(horizontal="center", vertical="center")
        ws.cell(row=start_row-1, column=col).font = opyxl.styles.Font(bold=True)
        

        # Week day
        ws.cell(row=start_row, column=col).value = datetime(next_month.year, next_month.month, day).strftime('%a') #short weekday name 
        ws.cell(row=start_row, column=col).alignment = opyxl.styles.Alignment(horizontal="center", vertical="center")
        ws.cell(row=start_row, column=col).font = opyxl.styles.Font(bold=True)
        
        # Color weekend and holidays
        date = f"{next_month.year}-{next_month.month:02d}-{day:02d}"
        for n_time_slot in range(n_rows + 1):
            if calendar.day_name[calendar.weekday(next_month.year, next_month.month, day)] in weekly_day_off or date in holidays.US(years=next_month.year):
                ws.cell(row=start_row+n_time_slot, column=col).fill = opyxl.styles.PatternFill(start_color=color_holidays, end_color=color_holidays, fill_type="solid")
            else:
                ws.cell(row=start_row+n_time_slot, column=col).fill = opyxl.styles.PatternFill(start_color=color_workday, end_color=color_workday, fill_type="solid")

        # Unlock cells to be modified
        for n_time_slot in range(n_rows):
            ws.cell(row=start_row+n_time_slot+1, column=col).protection = opyxl.styles.Protection(locked=False)

    # Conditional formatting
    red_fill = opyxl.styles.PatternFill(start_color=color_holidays, end_color=color_holidays, fill_type="solid")
    gray_fill = opyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    rule_red = opyxl.formatting.rule.CellIsRule(operator='equal', formula=['"X"'], fill=red_fill)
    first_cell = f"{opyxl.utils.get_column_letter(start_col+1)}{start_row+1}"
    rule_gray = opyxl.formatting.rule.FormulaRule(formula=[f'AND({first_cell}<>"X", {first_cell}<>"")'], fill=gray_fill)
    ws.conditional_formatting.add(f"{first_cell}:{opyxl.utils.get_column_letter(days_in_month + start_col)}{n_rows + start_row}", rule_red)
    ws.conditional_formatting.add(f"{first_cell}:{opyxl.utils.get_column_letter(days_in_month + start_col)}{n_rows + start_row}", rule_gray)

    # Borders
    thin_side = opyxl.styles.Side(border_style="thin", color="000000")
    border = opyxl.styles.Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for row in ws[f"{opyxl.utils.get_column_letter(start_col)}{start_row-1}:{opyxl.utils.get_column_letter(days_in_month + start_col)}{n_rows + start_row}"]:
        for cell in row:
            cell.border = border


    ws.protection.sheet = True
    ws.protection.password = 'passwd'
    pass       



def read_attendees(filename="attendees.xlsx"):
    wb = opyxl.load_workbook(filename)
    ws = wb.active

    attendees = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            attendees.append(cell.value)
    return attendees



def create_timeslots(slot_ranges, slot_duration):

    time_slots = []
    for start_hour, end_hour in slot_ranges:
        current_time = start_hour
        while current_time < end_hour:
            hours = int(current_time)
            minutes = int((current_time - hours) * 60)
            time_slots.append(f"{hours:02d}:{minutes:02d}")
            current_time += slot_duration

    return time_slots